from flask import Blueprint, render_template, redirect, url_for, flash, request, abort, send_file
from flask_login import login_required, current_user
from models import User, Asset, Role, Project
from extensions import db
from email_utils import send_welcome_email
from functools import wraps
import secrets
import io

users_bp = Blueprint('users', __name__)

DEFAULT_ROLE_CHOICES = [
    ('admin', 'Administrator', 'Full system administrator access.'),
    ('finance', 'Finance Officer', 'Finance team access and reporting approvals.'),
    ('user', 'Staff User', 'Standard staff access to own assets and requests.'),
]


def _roles():
    roles = Role.query.order_by(Role.name).all()
    if roles:
        return roles

    defaults = [
        {'name': 'admin', 'label': 'Administrator', 'description': 'Full system administrator access.'},
        {'name': 'finance', 'label': 'Finance Officer', 'description': 'Finance team access and reporting approvals.'},
        {'name': 'user', 'label': 'Staff User', 'description': 'Standard staff access to own assets and requests.'},
    ]
    created = False
    for data in defaults:
        if not Role.query.filter_by(name=data['name']).first():
            db.session.add(Role(**data))
            created = True
    if created:
        db.session.commit()
    return Role.query.order_by(Role.name).all()


def _countries():
    return Project.query.filter_by(status='active').order_by(Project.name).all()


def admin_required(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        if not current_user.is_authenticated or current_user.role != 'admin':
            abort(403)
        return f(*args, **kwargs)
    return decorated


@users_bp.route('/')
@login_required
@admin_required
def index():
    users = User.query.order_by(User.name).all()
    return render_template('users/index.html', users=users)


@users_bp.route('/import/template')
@login_required
@admin_required
def import_template():
    try:
        import openpyxl
    except ImportError:
        flash('openpyxl is not installed. Run: pip install openpyxl', 'error')
        return redirect(url_for('users.index'))

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Users'
    ws.append(['NAME', 'USERNAME', 'EMAIL', 'DEPARTMENT', 'COUNTRY', 'ROLE'])
    ws.append(['Jane Doe', 'janedoe', 'jane.doe@example.com', 'Finance', 'Uganda', 'finance'])

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    return send_file(
        buffer,
        download_name='user_import_template.xlsx',
        as_attachment=True,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )


@users_bp.route('/import', methods=['POST'])
@login_required
@admin_required
def import_users():
    f = request.files.get('file')
    if not f or not f.filename.lower().endswith('.xlsx'):
        flash('Please upload a valid .xlsx Excel file.', 'error')
        return redirect(url_for('users.index'))

    try:
        import openpyxl
    except ImportError:
        flash('openpyxl is not installed. Run: pip install openpyxl', 'error')
        return redirect(url_for('users.index'))

    try:
        wb = openpyxl.load_workbook(f, data_only=True)
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))
    except Exception as e:
        flash(f'Could not read Excel file: {e}', 'error')
        return redirect(url_for('users.index'))

    if len(rows) < 2:
        flash('The file is empty or has no data rows.', 'error')
        return redirect(url_for('users.index'))

    headers = [str(h).strip().lower().replace(' ', '_') if h is not None else '' for h in rows[0]]

    def col(row, name):
        if name in headers:
            v = row[headers.index(name)]
            return str(v).strip() if v is not None else ''
        return ''

    role_map = {r.name.lower(): r.name for r in Role.query.all()}
    role_label_map = {r.label.lower(): r.name for r in Role.query.all()}
    proj_map = {p.name.lower(): p for p in Project.query.all()}
    proj_code = {p.code.lower(): p for p in Project.query.all()}

    imported = 0
    skipped = 0
    errors = []

    for i, row in enumerate(rows[1:], start=2):
        if all(v is None or str(v).strip() == '' for v in row):
            continue

        name = col(row, 'name')
        username = col(row, 'username')
        email = col(row, 'email')
        department = col(row, 'department')
        country_value = col(row, 'country')
        role_value = col(row, 'role')

        if not name or not username:
            errors.append(f'Row {i}: NAME and USERNAME are required. Row skipped.')
            skipped += 1
            continue

        username = username.lower()
        if User.query.filter_by(username=username).first():
            errors.append(f'Row {i}: Username {username} already exists. Row skipped.')
            skipped += 1
            continue

        role_key = 'user'
        if role_value:
            requested = role_value.strip().lower()
            role_key = role_map.get(requested) or role_label_map.get(requested) or 'user'
            if requested and role_key == 'user' and requested not in ('user', 'staff', 'staff user'):
                errors.append(f'Row {i}: Role {role_value} not found; defaulted to user.')

        project = None
        if country_value:
            requested_country = country_value.strip().lower()
            project = proj_map.get(requested_country) or proj_code.get(requested_country)
            if not project:
                errors.append(f'Row {i}: Country {country_value} not found; left blank.')

        password = ''.join(secrets.choice('abcdefghijklmnopqrstuvwxyzABCDEFGHIJKLMNOPQRSTUVWXYZ0123456789') for _ in range(10))
        user = User(
            username=username,
            name=name,
            email=email or None,
            department=department or None,
            project_id=project.id if project else None,
            role=role_key,
            status='active'
        )
        user.set_password(password)
        db.session.add(user)
        imported += 1

    if imported:
        db.session.commit()

    users = User.query.order_by(User.name).all()
    summary = {
        'imported': imported,
        'skipped': skipped,
        'errors': errors,
    }
    if imported:
        flash(f'Imported {imported} user(s). {skipped} row(s) skipped.', 'success')
    elif skipped:
        flash('No users were imported. Please check the file and try again.', 'error')

    return render_template('users/index.html', users=users, import_summary=summary)


@users_bp.route('/new', methods=['GET', 'POST'])
@login_required
@admin_required
def new():
    if request.method == 'POST':
        username = request.form['username'].strip()
        password = request.form['password']

        if User.query.filter_by(username=username).first():
            flash('Username already exists.', 'error')
            return render_template('users/form.html', user=None, title='Register User', roles=_roles(), countries=_countries())

        project_id = request.form.get('project_id', '').strip()
        user = User(
            username   = username,
            name       = request.form['name'].strip(),
            email      = request.form.get('email', '').strip(),
            department = request.form.get('department', '').strip(),
            project_id = int(project_id) if project_id else None,
            role       = request.form.get('role', 'user'),
            status     = request.form.get('status', 'active'),
        )
        user.set_password(password)
        db.session.add(user)
        db.session.commit()

        # ── Send welcome email ────────────────────────────────────
        send_email = request.form.get('send_email') == 'on'
        if send_email and user.email:
            ok, err = send_welcome_email(
                to_email = user.email,
                to_name  = user.name,
                username = user.username,
                role     = user.role
            )
            if ok:
                flash(f'✓ User {user.name} registered and account notification sent to {user.email}.', 'success')
            else:
                flash(f'✓ User {user.name} registered but email failed: {err}', 'warning')
        elif send_email and not user.email:
            flash(f'✓ User {user.name} registered. Email not sent — no email address provided.', 'warning')
        else:
            flash(f'✓ User {user.name} registered successfully.', 'success')

        return redirect(url_for('users.index'))

    return render_template('users/form.html', user=None, title='Register New User', roles=_roles(), countries=_countries())


@users_bp.route('/roles')
@login_required
@admin_required
def roles():
    return render_template('users/roles.html', roles=_roles(), title='Manage Roles')


@users_bp.route('/roles/new', methods=['POST'])
@login_required
@admin_required
def role_new():
    name = request.form['name'].strip().lower()
    label = request.form['label'].strip()
    description = request.form.get('description', '').strip()
    if not name or not label:
        flash('Role key and display name are required.', 'error')
        return redirect(url_for('users.roles'))
    if Role.query.filter_by(name=name).first():
        flash('Role key already exists.', 'error')
        return redirect(url_for('users.roles'))
    role = Role(name=name, label=label, description=description)
    db.session.add(role)
    db.session.commit()
    flash('Role added.', 'success')
    return redirect(url_for('users.roles'))


@users_bp.route('/roles/<int:role_id>/edit', methods=['POST'])
@login_required
@admin_required
def role_edit(role_id):
    role = Role.query.get_or_404(role_id)
    new_name = request.form['name'].strip().lower()
    new_label = request.form['label'].strip()
    description = request.form.get('description', '').strip()
    if not new_name or not new_label:
        flash('Role key and display name are required.', 'error')
        return redirect(url_for('users.roles'))
    dup = Role.query.filter_by(name=new_name).first()
    if dup and dup.id != role_id:
        flash('Role key already exists.', 'error')
        return redirect(url_for('users.roles'))
    role.name = new_name
    role.label = new_label
    role.description = description
    db.session.commit()
    flash('Role updated.', 'success')
    return redirect(url_for('users.roles'))


@users_bp.route('/roles/<int:role_id>/delete', methods=['POST'])
@login_required
@admin_required
def role_delete(role_id):
    role = Role.query.get_or_404(role_id)
    if User.query.filter_by(role=role.name).first():
        flash('Cannot delete: one or more users are assigned this role.', 'error')
        return redirect(url_for('users.roles'))
    db.session.delete(role)
    db.session.commit()
    flash('Role deleted.', 'success')
    return redirect(url_for('users.roles'))


@users_bp.route('/<int:user_id>/edit', methods=['GET', 'POST'])
@login_required
@admin_required
def edit(user_id):
    user = User.query.get_or_404(user_id)

    if request.method == 'POST':
        new_username = request.form['username'].strip()
        dup = User.query.filter_by(username=new_username).first()
        if dup and dup.id != user_id:
            flash('Username already taken.', 'error')
            return render_template('users/form.html', user=user, title='Edit User', roles=_roles(), countries=_countries())

        project_id = request.form.get('project_id', '').strip()
        user.username   = new_username
        user.name       = request.form['name'].strip()
        user.email      = request.form.get('email', '').strip()
        user.department = request.form.get('department', '').strip()
        user.project_id = int(project_id) if project_id else None
        user.role       = request.form.get('role', user.role)
        user.status     = request.form.get('status', user.status)

        pw = request.form.get('password', '').strip()
        if pw:
            user.set_password(pw)

        db.session.commit()
        flash(f'✓ {user.name} updated successfully.', 'success')
        return redirect(url_for('users.index'))

    return render_template('users/form.html', user=user, title='Edit User', roles=_roles(), countries=_countries())


@users_bp.route('/<int:user_id>/delete', methods=['POST'])
@login_required
@admin_required
def delete(user_id):
    user = User.query.get_or_404(user_id)
    if user.id == current_user.id:
        flash('You cannot delete your own account.', 'error')
        return redirect(url_for('users.index'))
    if Asset.query.filter_by(assigned_to_id=user_id).first():
        flash('Cannot delete: user has assets assigned. Unassign assets first.', 'error')
        return redirect(url_for('users.index'))
    db.session.delete(user)
    db.session.commit()
    flash('User deleted.', 'success')
    return redirect(url_for('users.index'))
