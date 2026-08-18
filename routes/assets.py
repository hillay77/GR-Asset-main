from flask import (Blueprint, render_template, redirect, url_for,
                   flash, request, abort, make_response, send_file)
from flask_login import login_required, current_user
from models import Asset, AssetCategory, Project, Vendor, User, ReturnRecord
from extensions import db
from datetime import date, datetime
from functools import wraps
import csv
import io
import os
from pdf_utils import build_gr_pdf, build_pdf_table, asset_register_col_widths, pdf_response

assets_bp = Blueprint('assets', __name__)


def admin_required(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        if not current_user.is_authenticated or current_user.role != 'admin':
            abort(403)
        return f(*args, **kwargs)
    return decorated


def admin_or_finance(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        if not current_user.is_authenticated or current_user.role not in ('admin', 'finance'):
            abort(403)
        return f(*args, **kwargs)
    return decorated


def _next_asset_number():
    last = db.session.query(db.func.max(Asset.asset_number)).scalar()
    return (last or 0) + 1


def _build_tag(project_code, asset_number, category_code):
    return f"GR-{project_code}-{category_code}-{str(asset_number).zfill(4)}"


def _format_age(asset):
    if asset.age_years is not None or asset.age_months is not None:
        y = asset.age_years or 0
        m = asset.age_months or 0
        parts = []
        if y: parts.append(f"{y} yr")
        if m: parts.append(f"{m} mo")
        return ' '.join(parts) if parts else '—'
    if not asset.date_purchased:
        return '—'
    today = datetime.utcnow().date()
    dp = asset.date_purchased
    years = today.year - dp.year - ((today.month, today.day) < (dp.month, dp.day))
    months = (today.month - dp.month - (today.day < dp.day)) % 12
    parts = []
    if years: parts.append(f"{years} yr")
    if months: parts.append(f"{months} mo")
    return ' '.join(parts) if parts else '0 mo'


def _asset_query():
    q = request.args.get('q', '').strip()
    proj_id = request.args.get('project_id', type=int)
    cond = request.args.get('condition', '').strip()

    query = Asset.query.filter_by(status='active')
    if current_user.is_staff_portal:
        query = query.filter_by(assigned_to_id=current_user.id)
    if q:
        query = query.filter(
            db.or_(Asset.name.ilike(f'%{q}%'),
                   Asset.tag.ilike(f'%{q}%'),
                   Asset.serial_number.ilike(f'%{q}%'))
        )
    if proj_id:
        query = query.filter_by(project_id=proj_id)
    if cond:
        query = query.filter_by(condition=cond)
    return query.order_by(Asset.asset_number)


def _build_asset_rows(assets):
    rows = [[
        'Asset Tag', 'Category', 'Name', 'Processor', 'Serial Number', 'Department',
        'Date Purchased', 'Age', 'Assigned To', 'Condition', 'Vendor'
    ]]
    for a in assets:
        rows.append([
            a.tag,
            a.category.name if a.category else '—',
            a.name,
            a.processor or '—',
            a.serial_number or '—',
            a.assigned_user.department if a.assigned_user and a.assigned_user.department else '—',
            a.date_purchased.strftime('%d %b %Y') if a.date_purchased else '—',
            _format_age(a),
            a.assigned_user.name if a.assigned_user else '—',
            a.condition.title(),
            a.vendor.name if a.vendor else '—'
        ])
    return rows


def _get_or_create_default_project():
    """Return the Project instance for the default 'GEN' project, creating it if needed."""
    proj = Project.query.filter_by(code='GEN').first()
    if proj:
        return proj
    proj = Project(code='GEN', name='General', description='Default project', status='active')
    db.session.add(proj)
    try:
        db.session.commit()
    except Exception:
        db.session.rollback()
        proj = Project.query.filter_by(code='GEN').first()
    return proj


# ── List all assets ────────────────────────────────────────────────────────
@assets_bp.route('/')
@login_required
def index():
    q         = request.args.get('q', '').strip()
    proj_id   = request.args.get('project_id', type=int)
    cond      = request.args.get('condition', '').strip()

    query = _asset_query()
    assets = query.all()
    projects = Project.query.filter_by(status='active').all()
    return render_template('assets/index.html',
                           assets=assets, projects=projects,
                           q=q, proj_id=proj_id, cond=cond)


@assets_bp.route('/download/csv')
@login_required
def download_csv():
    assets = _asset_query().all()
    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow(['GR Asset Management System - Asset Register'])
    writer.writerows(_build_asset_rows(assets))
    csv_data = output.getvalue()
    output.close()

    response = make_response(csv_data)
    response.headers['Content-Type'] = 'text/csv; charset=utf-8'
    response.headers['Content-Disposition'] = 'attachment; filename=gr_assets.csv'
    return response


@assets_bp.route('/download/pdf')
@login_required
def download_pdf():
    assets = _asset_query().all()

    def body(doc):
        return [build_pdf_table(_build_asset_rows(assets), asset_register_col_widths(doc.width))]

    return pdf_response(build_gr_pdf(body), 'gr_assets.pdf')


# ── Register asset ─────────────────────────────────────────────────────────
@assets_bp.route('/new', methods=['GET', 'POST'])
@login_required
@admin_required
def new():
    categories = AssetCategory.query.order_by(AssetCategory.name).all()
    projects   = Project.query.filter_by(status='active').order_by(Project.name).all()
    vendors    = Vendor.query.order_by(Vendor.name).all()
    users      = User.query.filter_by(status='active').order_by(User.name).all()

    if request.method == 'POST':
        cat_id      = request.form.get('category_id', type=int)
        proj_id     = request.form.get('project_id',  type=int)
        if not proj_id:
            flash('Please select a country.', 'error')
            return render_template('assets/form.html',
                                   categories=categories, projects=projects,
                                   vendors=vendors, users=users,
                                   asset=None, title='Register New Asset')
        cat         = AssetCategory.query.get_or_404(cat_id)
        proj        = Project.query.get_or_404(proj_id)

        num = _next_asset_number()
        proj_code = proj.code
        tag = _build_tag(proj_code, num, cat.code)

        assign_id = request.form.get('assigned_to_id', type=int) or None
        dp = request.form.get('date_purchased')

        processor = request.form.get('processor', '').strip() or None
        age_years = request.form.get('age_years', type=int) or None
        age_months = request.form.get('age_months', type=int) or None

        asset = Asset(
            asset_number   = num,
            tag            = tag,
            name           = request.form['name'].strip(),
            serial_number  = request.form.get('serial_number', '').strip(),
            description    = request.form.get('description', '').strip(),
            price          = request.form.get('price') or 0,
            date_purchased = date.fromisoformat(dp) if dp else None,
            condition      = request.form.get('condition', 'good'),
            location       = request.form.get('location', '').strip(),
            category_id    = cat_id,
            project_id     = proj.id,
            vendor_id      = request.form.get('vendor_id', type=int) or None,
            assigned_to_id = assign_id,
            assigned_on    = date.today() if assign_id else None,
            status         = 'active',
            processor      = processor,
            age_years      = age_years,
            age_months     = age_months,
        )
        db.session.add(asset)
        db.session.commit()
        flash(f'Asset {tag} registered successfully.', 'success')
        return redirect(url_for('assets.index'))

    return render_template('assets/form.html',
                           categories=categories, projects=projects,
                           vendors=vendors, users=users,
                           asset=None, title='Register New Asset')


# ── Edit asset ─────────────────────────────────────────────────────────────
@assets_bp.route('/<int:asset_id>/edit', methods=['GET', 'POST'])
@login_required
@admin_required
def edit(asset_id):
    asset      = Asset.query.get_or_404(asset_id)
    categories = AssetCategory.query.order_by(AssetCategory.name).all()
    projects   = Project.query.filter_by(status='active').order_by(Project.name).all()
    vendors    = Vendor.query.order_by(Vendor.name).all()
    users      = User.query.filter_by(status='active').order_by(User.name).all()

    if request.method == 'POST':
        cat_id      = request.form.get('category_id', type=int)
        proj_id     = request.form.get('project_id',  type=int)
        if not proj_id:
            flash('Please select a country.', 'error')
            return render_template('assets/form.html',
                                   categories=categories, projects=projects,
                                   vendors=vendors, users=users,
                                   asset=asset, title='Edit Asset')
        cat         = AssetCategory.query.get_or_404(cat_id)
        proj        = Project.query.get_or_404(proj_id)

        new_assign = request.form.get('assigned_to_id', type=int) or None
        dp = request.form.get('date_purchased')

        asset.name           = request.form['name'].strip()
        asset.serial_number  = request.form.get('serial_number', '').strip()
        asset.description    = request.form.get('description', '').strip()
        asset.price          = request.form.get('price') or 0
        asset.date_purchased = date.fromisoformat(dp) if dp else None
        asset.condition      = request.form.get('condition', 'good')
        asset.location       = request.form.get('location', '').strip()
        asset.category_id    = cat_id
        asset.project_id     = proj.id
        asset.vendor_id      = request.form.get('vendor_id', type=int) or None
        asset.processor      = request.form.get('processor', '').strip() or None
        asset.age_years      = request.form.get('age_years', type=int) or None
        asset.age_months     = request.form.get('age_months', type=int) or None
        asset.updated_at     = datetime.utcnow()

        # Rebuild tag from country + category
        asset.tag = _build_tag(proj.code, asset.asset_number, cat.code)

        # Track assignment date change
        if new_assign != asset.assigned_to_id:
            asset.assigned_to_id = new_assign
            asset.assigned_on    = date.today() if new_assign else None

        db.session.commit()
        flash(f'Asset {asset.tag} updated.', 'success')
        return redirect(url_for('assets.index'))

    return render_template('assets/form.html',
                           categories=categories, projects=projects,
                           vendors=vendors, users=users,
                           asset=asset, title='Edit Asset')


# ── Delete asset ───────────────────────────────────────────────────────────
@assets_bp.route('/<int:asset_id>/delete', methods=['POST'])
@login_required
@admin_required
def delete(asset_id):
    asset = Asset.query.get_or_404(asset_id)
    db.session.delete(asset)
    db.session.commit()
    flash('Asset deleted.', 'success')
    return redirect(url_for('assets.index'))


# ── Assign page ────────────────────────────────────────────────────────────
@assets_bp.route('/assign')
@login_required
@admin_required
def assign():
    users = User.query.filter(
        User.status == 'active',
        User.role != 'admin'
    ).order_by(User.name).all()
    return render_template('assets/assign.html', users=users, selected_user=None)


@assets_bp.route('/assign/<int:user_id>')
@login_required
@admin_required
def assign_user(user_id):
    user      = User.query.get_or_404(user_id)
    users     = User.query.filter(User.status == 'active', User.role != 'admin').order_by(User.name).all()
    my_assets = Asset.query.filter_by(assigned_to_id=user_id, status='active').all()
    available = Asset.query.filter_by(assigned_to_id=None, status='active').all()
    return render_template('assets/assign.html',
                           users=users,
                           selected_user=user,
                           my_assets=my_assets,
                           available=available)


@assets_bp.route('/assign/<int:user_id>/do', methods=['POST'])
@login_required
@admin_required
def do_assign(user_id):
    asset_id = request.form.get('asset_id', type=int)
    asset = Asset.query.get_or_404(asset_id)
    asset.assigned_to_id = user_id
    asset.assigned_on    = date.today()
    db.session.commit()
    flash(f'{asset.tag} assigned successfully.', 'success')
    return redirect(url_for('assets.assign_user', user_id=user_id))


@assets_bp.route('/unassign/<int:asset_id>', methods=['POST'])
@login_required
@admin_required
def unassign(asset_id):
    asset = Asset.query.get_or_404(asset_id)
    user_id = asset.assigned_to_id
    asset.assigned_to_id = None
    asset.assigned_on    = None
    db.session.commit()
    flash(f'{asset.tag} unassigned.', 'success')
    return redirect(url_for('assets.assign_user', user_id=user_id))


# ── Return asset ───────────────────────────────────────────────────────────
@assets_bp.route('/<int:asset_id>/return', methods=['GET', 'POST'])
@login_required
def return_asset(asset_id):
    """
    Staff users are redirected to the approval-based return request flow.
    Admin can still process a direct return from the assign page via unassign.
    """
    asset = Asset.query.get_or_404(asset_id)
    # Always redirect staff to the request-based return flow
    if current_user.is_staff_portal:
        return redirect(url_for('requests.new_return_request', asset_id=asset_id))
    # Admin direct return (from assign page unassign button) — keep as-is
    abort(403)


# ── Print form ─────────────────────────────────────────────────────────────
@assets_bp.route('/<int:asset_id>')
@login_required
def detail(asset_id):
    asset = Asset.query.get_or_404(asset_id)
    if current_user.is_staff_portal and asset.assigned_to_id != current_user.id:
        abort(403)
    return render_template('assets/detail.html', asset=asset)


@assets_bp.route('/<int:asset_id>/receipt.pdf')
@login_required
def asset_receipt_pdf(asset_id):
    from pdf_utils import build_gr_pdf, build_pdf_table, scaled_col_widths, pdf_response
    from reportlab.platypus import Paragraph, Spacer
    from reportlab.lib.styles import getSampleStyleSheet

    asset = Asset.query.get_or_404(asset_id)
    if current_user.is_staff_portal and asset.assigned_to_id != current_user.id:
        abort(403)
    user = asset.assigned_user or current_user

    rows = [
        ['Field', 'Value'],
        ['Asset Tag', asset.tag],
        ['Asset Name', asset.name],
        ['Category', asset.category.name if asset.category else '—'],
        ['Country', asset.project.name if asset.project else '—'],
        ['Serial Number', asset.serial_number or '—'],
        ['Condition', asset.condition.title()],
        ['Assigned To', user.name],
        ['Department', user.department or '—'],
        ['Assigned On', asset.assigned_on.strftime('%d %b %Y') if asset.assigned_on else '—'],
        ['Purchase Value', f'USD {float(asset.price or 0):,.2f}'],
    ]

    def body(doc):
        styles = getSampleStyleSheet()
        title = Paragraph(f'Asset Receipt — {asset.tag}', styles['Heading2'])
        return [
            title,
            Spacer(1, 8),
            build_pdf_table(rows, scaled_col_widths([80, 200], doc.width)),
        ]

    return pdf_response(build_gr_pdf(body), f'receipt_{asset.tag}.pdf')


@assets_bp.route('/<int:asset_id>/qrcode.pdf')
@login_required
def asset_qrcode_pdf(asset_id):
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.platypus import SimpleDocTemplate

    asset = Asset.query.get_or_404(asset_id)
    if current_user.is_staff_portal and asset.assigned_to_id != current_user.id:
        abort(403)

    page_width, page_height = landscape(A4)
    label_width = 260
    label_height = 80
    qr_size = 64

    buffer = io.BytesIO()
    doc = SimpleDocTemplate(
        buffer,
        pagesize=landscape(A4),
        leftMargin=28,
        rightMargin=28,
        topMargin=28,
        bottomMargin=28,
    )

    label = _build_asset_qr_label(asset, label_width, label_height, qr_size)
    doc.build([label])
    pdf_bytes = buffer.getvalue()
    buffer.close()
    return pdf_response(pdf_bytes, f'{asset.tag}_qr.pdf')


@assets_bp.route('/qrcodes/pdf')
@login_required
def asset_qrcodes_pdf():
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Spacer
    from reportlab.lib import colors

    assets = _asset_query().all()

    page_width, page_height = landscape(A4)
    gutter = 8
    label_width = 260
    label_height = 80
    qr_size = 64

    labels = [_build_asset_qr_label(asset, label_width, label_height, qr_size) for asset in assets]
    rows = []
    row_heights = []
    for i in range(0, len(labels), 2):
        first = labels[i]
        second = labels[i + 1] if i + 1 < len(labels) else Spacer(label_width, label_height)
        rows.append([first, Spacer(gutter, label_height), second])
        row_heights.append(label_height)
        if i + 2 < len(labels):
            rows.append([Spacer(label_width, 12), Spacer(gutter, 12), Spacer(label_width, 12)])
            row_heights.append(12)

    table = Table(rows, colWidths=[label_width, gutter, label_width], rowHeights=row_heights, hAlign='LEFT')
    table.setStyle(TableStyle([
        ('LEFTPADDING', (0, 0), (-1, -1), 0),
        ('RIGHTPADDING', (0, 0), (-1, -1), 0),
        ('TOPPADDING', (0, 0), (-1, -1), 0),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 0),
    ]))

    buffer = io.BytesIO()
    doc = SimpleDocTemplate(
        buffer,
        pagesize=landscape(A4),
        leftMargin=28,
        rightMargin=28,
        topMargin=28,
        bottomMargin=28,
    )
    doc.build([table])
    pdf_bytes = buffer.getvalue()
    buffer.close()
    return pdf_response(pdf_bytes, 'gr_asset_qrcodes.pdf')


def _asset_qr_payload(asset):
    return '\n'.join([
        f'GR Asset Tag: {asset.tag}',
        f'Name: {asset.name}',
        f'Category: {asset.category.name if asset.category else "—"}',
        f'Country: {asset.project.name if asset.project else "—"}',
        f'Location: {asset.location or "—"}',
        f'Serial Number: {asset.serial_number or "—"}',
        f'Condition: {asset.condition.title()}',
        f'Assigned To: {asset.assigned_user.name if asset.assigned_user else "—"}',
        f'Assigned On: {asset.assigned_on.strftime("%d %b %Y") if asset.assigned_on else "—"}',
    ])


def _load_logo_image(logo_path, width, height):
    from reportlab.platypus import Image as RLImage, Spacer as RLSpacer

    if not os.path.isfile(logo_path):
        return RLSpacer(width, height)

    try:
        return RLImage(logo_path, width=width, height=height)
    except Exception:
        return RLSpacer(width, height)


def _build_asset_qr_label(asset, label_width, label_height, qr_size):
    from reportlab.graphics.barcode.qr import QrCodeWidget
    from reportlab.graphics.shapes import Drawing
    from reportlab.platypus import Image, Paragraph, Spacer, Table, TableStyle
    from reportlab.lib.styles import ParagraphStyle
    from reportlab.lib import colors

    qr = QrCodeWidget(_asset_qr_payload(asset), barWidth=qr_size, barHeight=qr_size, x=0, y=0)
    qr_group = qr.draw()
    qr_drawing = Drawing(qr_size, qr_size)
    qr_drawing.add(qr_group)

    heading_style = ParagraphStyle('QrLabelHeading', fontName='Helvetica-Bold', fontSize=12, textColor=colors.HexColor('#0f172a'), leading=14, alignment=1)
    tag_style = ParagraphStyle('QrLabelTag', fontName='Helvetica', fontSize=12, textColor=colors.HexColor('#0f172a'), leading=14, alignment=1)
    detail_style = ParagraphStyle('QrLabelDetail', fontName='Helvetica', fontSize=12, textColor=colors.HexColor('#0f172a'), leading=14, alignment=1)
    note_style = ParagraphStyle('QrLabelNote', fontName='Helvetica', fontSize=7, textColor=colors.HexColor('#475569'), leading=9, alignment=1)

    center_width = label_width - qr_size - 54
    center_content = Table([
        [Paragraph('GR ASSET TAG', heading_style)],
        [Spacer(1, 1)],
        [Paragraph(asset.tag, tag_style)],
        [Spacer(1, 1)],
        [Paragraph(asset.name or '', detail_style)],
        [Spacer(1, 1)],
        [Paragraph('Scan to view asset details', note_style)],
    ], colWidths=[center_width - 4])
    center_content.setStyle(TableStyle([
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('LEFTPADDING', (0, 0), (-1, -1), 1),
        ('RIGHTPADDING', (0, 0), (-1, -1), 1),
        ('TOPPADDING', (0, 0), (-1, -1), 0),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 0),
    ]))

    center_block = Table([[center_content]], colWidths=[center_width], rowHeights=[qr_size])
    center_block.setStyle(TableStyle([
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('LEFTPADDING', (0, 0), (-1, -1), 0),
        ('RIGHTPADDING', (0, 0), (-1, -1), 0),
        ('TOPPADDING', (0, 0), (-1, -1), 0),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 0),
    ]))

    static_img_dir = os.path.abspath(os.path.join(os.path.dirname(__file__), '..', 'static', 'img'))
    logo_path = None
    for filename in ('GR.png', 'GR.jpg'):
        candidate = os.path.join(static_img_dir, filename)
        if os.path.isfile(candidate):
            logo_path = candidate
            break

    right_logo = _load_logo_image(logo_path or os.path.join(static_img_dir, 'GR.jpg'), width=48, height=48)

    logo_block = Table([
        [right_logo]
    ], colWidths=[54], rowHeights=[label_height])
    logo_block.setStyle(TableStyle([
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('LEFTPADDING', (0, 0), (-1, -1), 0),
        ('RIGHTPADDING', (0, 0), (-1, -1), 0),
        ('TOPPADDING', (0, 0), (-1, -1), 0),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 0),
    ]))

    qr_block = Table([[qr_drawing]], colWidths=[qr_size], rowHeights=[qr_size])
    qr_block.setStyle(TableStyle([
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('LEFTPADDING', (0, 0), (-1, -1), 0),
        ('RIGHTPADDING', (0, 0), (-1, -1), 0),
        ('TOPPADDING', (0, 0), (-1, -1), 0),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 0),
    ]))

    label = Table([
        [logo_block, center_block, qr_block]
    ], colWidths=[54, center_width, qr_size], rowHeights=[label_height])
    label.setStyle(TableStyle([
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('LEFTPADDING', (0, 0), (-1, -1), 1),
        ('RIGHTPADDING', (0, 0), (-1, -1), 1),
        ('TOPPADDING', (0, 0), (-1, -1), 1),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 1),
        ('BOX', (0, 0), (-1, -1), 0.75, colors.HexColor('#1a3a5c')),
    ]))
    return label


def build_asset_info_table(rows, width):
    from reportlab.platypus import Table, TableStyle, Paragraph
    from reportlab.lib.styles import getSampleStyleSheet
    from reportlab.lib import colors

    styles = getSampleStyleSheet()
    header_style = styles['Heading6']
    header_style.fontSize = 8
    body_style = styles['BodyText']
    body_style.fontSize = 7

    data = []
    for idx, row in enumerate(rows):
        label = Paragraph(str(row[0]), header_style if idx == 0 else body_style)
        value = Paragraph(str(row[1]), header_style if idx == 0 else body_style)
        data.append([label, value])

    table = Table(data, colWidths=[width * 0.35, width * 0.65])
    table.setStyle(TableStyle([
        ('GRID', (0, 0), (-1, -1), 0.25, colors.HexColor('#d1d5db')),
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#f3f4f6')),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('VALIGN', (0, 0), (-1, -1), 'TOP'),
        ('LEFTPADDING', (0, 0), (-1, -1), 4),
        ('RIGHTPADDING', (0, 0), (-1, -1), 4),
        ('TOPPADDING', (0, 0), (-1, -1), 4),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
    ]))
    return table


@assets_bp.route('/print/<int:asset_id>')
@login_required
def print_form(asset_id):
    """Single-asset print (admin use). Staff get all-assets form."""
    if current_user.is_staff_portal:
        # Redirect staff to their full custody form
        return redirect(url_for('assets.print_user_form', user_id=current_user.id))

    asset = Asset.query.get_or_404(asset_id)
    return render_template('assets/print_single.html', asset=asset)


@assets_bp.route('/print/user/<int:user_id>')
@login_required
def print_user_form(user_id):
    """Print all assets for a given user (staff see their own, admin sees any)."""
    if current_user.is_staff_portal and current_user.id != user_id:
        abort(403)
    user   = User.query.get_or_404(user_id)
    assets = Asset.query.filter_by(assigned_to_id=user_id, status='active')\
                        .order_by(Asset.asset_number).all()
    total  = sum(float(a.price or 0) for a in assets)
    return render_template('assets/print_user.html',
                           user=user, assets=assets, total=total)


# ─────────────────────────────────────────────
#  BULK IMPORT FROM EXCEL
# ─────────────────────────────────────────────

@assets_bp.route('/import/template')
@login_required
@admin_required
def download_import_template():
    try:
        import openpyxl
    except ImportError:
        flash('openpyxl is not installed. Run: pip install openpyxl', 'error')
        return redirect(url_for('assets.bulk_import'))

    project = Project.query.order_by(Project.name).first()
    category = AssetCategory.query.order_by(AssetCategory.name).first()
    vendor = Vendor.query.order_by(Vendor.name).first()
    user = User.query.filter_by(status='active').order_by(User.name).first()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Assets'
    ws.append([
        'tag', 'name', 'country', 'category', 'vendor', 'serial_number',
        'price', 'date_purchased', 'condition', 'location', 'assigned_to', 'description',
    ])
    ws.append([
        'GR-UG-LP-0001',
        'Dell Latitude 5420',
        project.name if project else 'Uganda',
        category.name if category else 'Laptop',
        vendor.name if vendor else '',
        'SN-DELL-001',
        1800000,
        '2024-01-15',
        'good',
        'Kampala Office',
        user.username if user else '',
        'Intel Core i5, 8GB RAM',
    ])

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    return send_file(
        buffer,
        download_name='asset_import_template.xlsx',
        as_attachment=True,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )


@assets_bp.route('/import', methods=['GET', 'POST'])
@login_required
@admin_required
def bulk_import():
    if request.method == 'GET':
        # Pass lookup data so template can show expected column values
        categories = AssetCategory.query.order_by(AssetCategory.name).all()
        projects   = Project.query.order_by(Project.name).all()
        vendors    = Vendor.query.order_by(Vendor.name).all()
        users      = User.query.filter_by(status='active').order_by(User.name).all()
        return render_template('assets/import.html',
                               categories=categories, projects=projects,
                               vendors=vendors, users=users)

    # ── POST — process uploaded file ──────────────────────────
    f = request.files.get('file')
    if not f or not f.filename.endswith('.xlsx'):
        flash('Please upload a valid .xlsx Excel file.', 'error')
        return redirect(url_for('assets.bulk_import'))

    try:
        import openpyxl
    except ImportError:
        flash('openpyxl is not installed. Run: pip install openpyxl', 'error')
        return redirect(url_for('assets.bulk_import'))

    try:
        wb  = openpyxl.load_workbook(f, data_only=True)
        ws  = wb.active
        rows = list(ws.iter_rows(values_only=True))
    except Exception as e:
        flash(f'Could not read Excel file: {e}', 'error')
        return redirect(url_for('assets.bulk_import'))

    if len(rows) < 2:
        flash('The file is empty or has no data rows.', 'error')
        return redirect(url_for('assets.bulk_import'))

    # Normalise header row
    headers = [str(h).strip().lower().replace(' ', '_') if h else '' for h in rows[0]]

    def col(row, name):
        """Get cell value by column header name, return stripped string or None."""
        if name in headers:
            v = row[headers.index(name)]
            return str(v).strip() if v is not None else None
        return None

    # Build lookup dicts (case-insensitive)
    cat_map  = {c.name.lower(): c for c in AssetCategory.query.all()}
    cat_code = {c.code.lower():  c for c in AssetCategory.query.all()}
    proj_map = {p.name.lower():  p for p in Project.query.all()}
    proj_code= {p.code.lower():  p for p in Project.query.all()}
    vend_map = {v.name.lower():  v for v in Vendor.query.all()}
    user_map = {u.username.lower(): u for u in User.query.all()}
    user_name= {u.name.lower():    u for u in User.query.all()}

    imported = 0
    skipped  = 0
    errors   = []

    for i, row in enumerate(rows[1:], start=2):
        # Skip completely empty rows
        if all(v is None or str(v).strip() == '' for v in row):
            continue

        tag  = col(row, 'tag') or col(row, 'asset_tag')
        name = col(row, 'name') or col(row, 'asset_name')

        if not tag:
            errors.append(f'Row {i}: Missing asset tag — row skipped.')
            skipped += 1
            continue
        if not name:
            errors.append(f'Row {i}: Missing asset name — row skipped.')
            skipped += 1
            continue

        # Skip if tag already exists
        if Asset.query.filter_by(tag=tag).first():
            errors.append(f'Row {i}: Tag <strong>{tag}</strong> already exists — skipped.')
            skipped += 1
            continue

        # Resolve category
        cat_val = col(row, 'category') or ''
        category = cat_map.get(cat_val.lower()) or cat_code.get(cat_val.lower())
        if not category:
            errors.append(f'Row {i} ({tag}): Category "{cat_val}" not found — will be left blank.')

        # Resolve project / country
        proj_val = col(row, 'country') or col(row, 'project') or col(row, 'project_code') or ''
        project  = proj_map.get(proj_val.lower()) or proj_code.get(proj_val.lower())
        if not project:
            errors.append(f'Row {i} ({tag}): Project "{proj_val}" not found — row skipped.')
            skipped += 1
            continue

        # Resolve vendor (optional)
        vend_val = col(row, 'vendor') or col(row, 'vendor_name') or ''
        vendor   = vend_map.get(vend_val.lower()) if vend_val else None

        # Resolve assigned user (optional)
        user_val  = col(row, 'assigned_to') or col(row, 'username') or ''
        assigned  = user_map.get(user_val.lower()) or user_name.get(user_val.lower()) if user_val else None

        # Parse price
        price_val = col(row, 'price') or col(row, 'purchase_price') or '0'
        try:
            price = float(str(price_val).replace(',', '').replace('UGX', '').replace('USD', '').strip() or 0)
        except ValueError:
            price = 0

        # Parse date
        date_val = col(row, 'date_purchased') or col(row, 'purchase_date') or ''
        date_purchased = None
        if date_val:
            for fmt in ('%Y-%m-%d', '%d/%m/%Y', '%d-%m-%Y', '%m/%d/%Y', '%d %b %Y', '%d-%b-%Y'):
                try:
                    date_purchased = datetime.strptime(date_val, fmt).date()
                    break
                except ValueError:
                    continue

        # Condition
        cond_val  = (col(row, 'condition') or 'good').lower()
        condition = cond_val if cond_val in ('new', 'good', 'fair', 'poor') else 'good'

        # Asset number — use next available
        max_num = db.session.query(db.func.max(Asset.asset_number)).scalar() or 0
        asset_number = max_num + 1

        asset = Asset(
            asset_number   = asset_number,
            tag            = tag,
            name           = name,
            serial_number  = col(row, 'serial_number') or col(row, 'serial'),
            description    = col(row, 'description') or col(row, 'specs'),
            location       = col(row, 'location'),
            price          = price,
            date_purchased = date_purchased,
            condition      = condition,
            status         = 'active',
            category_id    = category.id if category else None,
            project_id     = project.id,
            vendor_id      = vendor.id  if vendor   else None,
            assigned_to_id = assigned.id if assigned else None,
            assigned_on    = date.today() if assigned else None,
        )
        db.session.add(asset)
        imported += 1

    try:
        db.session.commit()
    except Exception as e:
        db.session.rollback()
        flash(f'Database error during import: {e}', 'error')
        return redirect(url_for('assets.bulk_import'))

    # Summary flash
    if imported:
        flash(f'✓ Successfully imported {imported} asset{"s" if imported != 1 else ""}.', 'success')
    if skipped:
        flash(f'{skipped} row{"s" if skipped != 1 else ""} skipped.', 'warning')

    return render_template('assets/import_result.html',
                           imported=imported, skipped=skipped, errors=errors)
